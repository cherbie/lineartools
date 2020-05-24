from openpyxl import Workbook
import os
import json

error_fieldnames = ['message', 'type', 'function']

class ErrorLogger:
	def __init__(self, dirname: str, filename: str = 'error_log', headers=list()):
		if not os.path.exists(dirname):
			print(f'Creating Directory: {dirname}')
			os.mkdir(dirname)
		
		self.filename = f'{dirname}/{filename}.xlsx'
		self.wb = Workbook() # optimised for large file writing
		self.errors = list() # all parsed data
		
		# -- Add to headers --
		headers.extend(error_fieldnames)
		self.headers = ErrorLogger.setHeaders(headers)
		self.empty_error = dict.fromkeys(self.headers)
		self.setWorksheet()

	def add(self, entry: dict):
		if entry is None:
			return None
	
		self.errors.append({**self.empty_error, **entry, 'toString': json.dumps(entry)})
		

	def setHeaders(fieldnames: list):
		'''
		@param fieldnames - iterable list of fieldnames
		'''
		headers = {}
		for itx, fieldname in enumerate(fieldnames):
			f =  fieldname.lower()
			headers[f] = itx+1
		return headers

	def setWorksheet(self):
		self.ws = self.wb.active
		self.ws.title = 'Errors'
		self.ws.sheet_properties.tabColor = 'FF69B4'

		for header, pos in self.headers.items():
			self.ws.cell(row=1, column=pos, value=header)
		self.currentRow = 2 # set the last completed row

	def write(self):
		'''
		@param entries - array of dictionary entries with key's containing the col headers
		'''
		for error in self.errors:
			for key, value in error.items():
				key = key.lower() # convert to lowercase
				self.ws.cell(row=self.currentRow, column=self.headers.get(key, len(self.headers)+1), value=value)
			self.currentRow += 1 # increment row
		return

	def printHeaders(self):
		print(self.headers.keys())

	def getFilename(self):
		return self.filename

	def save(self):
		'''
		Saves the excel file to file and then closes the file properly.
		'''
		return self.wb.save(self.filename)