from openpyxl import load_workbook
import os

class FileReader:
    def __init__(self, filename):
        '''
        @param filename - location of xlsx workbook file (Medrio Browse Data File)
        '''
        if not os.path.exists(filename):
            raise Exception('IOERROR: Could not find the input file')
        self.filename = filename
        self.wb = load_workbook(filename=filename, read_only=True) # instance workbook

    def getSheetnames(self):
        return self.wb.sheetnames

    def getWorksheet(self, sheetname=None):
        '''
        @raise KeyError - if worksheet does not exist
        '''
        if sheetname is None:
            return self.wb.active
        else:
            return self.wb[sheetname]


    def getSheetHeaders(self, sheetname=None):
        '''
        @raise KeyError - if worksheet does not exist
        @param sheetname - name of the sheet to retrieve headers from
        '''
        ws = self.getWorksheet(sheetname);
        toprow = ws.iter_rows(max_row=1) # tuple containing tuple's of Cell objects
        # print(toprow)
        headers = []
        for row in toprow:
            for cell in row:
                # print(cell.value)
                if cell.value == None:
                    break
                else:
                    headers.append(cell.value)
            break
        return headers