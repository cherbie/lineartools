from openpyxl import load_workbook
import os

class FileReader:
    def __init__(self, filename):
        '''
        @param filename - location of xlsx workbook file
        '''
        if not os.path.exists(filename):
            raise Exception('IOERROR: Could not find the input file')
        self.filename = filename
        self.wb = load_workbook(filename=filename)

    def getSheetnames(self):
        return self.wb.sheetnames

    def getWorksheet(self, sheetname):
        '''
        @raise KeyError - if worksheet does not exist
        '''
        return self.wb[sheetname]


    def getSheetHeaders(self, sheetname):
        '''
        @raise KeyError - if worksheet does not exist
        @param sheetname - name of the sheet to retrieve headers from
        '''
        sheet = self.wb[sheetname]
        toprow = sheet['A:Z'] # tuple containing tuple's of Cell objects
        headers = []
        for header in toprow:
            print(header[0].value)
            if header[0].value == None:
                break
            else:
                headers.append(header[0].value)
        return headers
