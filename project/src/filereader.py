from openpyxl import load_workbook
import os

class FileReader:
    def __init__(self, filename):
        '''
        @param filename - location of xlsx workbook file
        '''
        if not os.path.exists(filename):
            raise IOError
        self.filename = filename
        self.wb = load_workbook(filename=filename)
    
    def getSheetnames(self):
        return self.wb.sheetnames
    
    def getWorksheet(self, sheetname):
        '''
        @raise KeyError - if worksheet does not exist
        '''
        return self.wb[sheetname]

    
    def getSheetHeaders(self, sheetname):
        '''
        @raise KeyError - if worksheet does not exist
        '''
        sheet = self.wb[sheetname]
        toprow = sheet['A:A'] # tuple containing tuple's of Cell objects
        headers = []
        for header in toprow:
            if header[0].value == None:
                break
            else:
                headers.append(header[0].value)
        return headers
        
        