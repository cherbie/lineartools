import sys, os
from openpyxl import Workbook, load_workbook
from filewriter import FileWriter
from filereader import FileReader
from parser import workbookParser

wb = Workbook()
# wb_ecg = load_workbook('test.xlsx')

visit_names = ('Screening', 'Day -2 [Period 1]', 'Day -1 [Period 1]', 'Day 1 [Period 1]', 'Day 2 [Period 1]', 'Day 3 [Period 1]',
'Day 4 [Period 1]', 'Day -2 [Period 2]', 'Day -1 [Period 2]', 'Day 1 [Period 2]', 'Day 2 [Period 2]', 'Day 3 [Period 2]',
'Day 4 [Period 2]', 'Day 7 [Period 2]  EOS', 'Unscheduled Visits')

ecg_names = ('12 Lead ECG (Triplicate)', 'Unscheduled 12-Lead Triplicate ECG', '45MIN Predose 12 Lead ECG (Triplicate)',
'30MIN Predose 12 Lead ECG (Triplicate)', '15MIN Predose 12 Lead ECG (Triplicate)', '30MIN 12 Lead ECG (Triplicate)',
'1HR 12 Lead ECG (Triplicate)', '1.5HR 12 Lead ECG (Triplicate)', '2HR 12 Lead ECG (Triplicate)', '3HR 12 Lead ECG (Triplicate)',
'4HR 12 Lead ECG (Triplicate)', '5HR 12 Lead ECG (Triplicate)', '6HR 12 Lead ECG (Triplicate)', '8HR 12 Lead ECG (Triplicate)',
'12HR 12 Lead ECG (Triplicate)', '16HR 12 Lead ECG (Triplicate)', '24HR 12 Lead ECG (Triplicate)', '48HR 12 Lead ECG (Triplicate)',
'72HR 12 Lead ECG (Triplicate)', '144HR 12 Lead ECG (Triplicate)')

print(sys.argv)

def main():
    # improve command line arguments handling
    # use a json config file
    template_filename = sys.argv[1]
    if os.path.exists(template_filename) :
        print(template_filename)
    else:
        raise IOError

    wb_class = FileReader(template_filename, {'visits': visit_names, 'forms': ecg_names})
    print(wb_class.getSheetnames())
    workbookParser(wb_class)

    
    # visit_names = input()
    # ecg_names  = input()
    
    print(visit_names)

main()