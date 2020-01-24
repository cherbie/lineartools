from openpyxl import Workbook
import os

class FileWriterXL:
    def __init__(self, name, fieldnames):
        dirname = os.path.dirname(name)
        if not os.path.exists(dirname):
            print(f'creating directory: {dirname}')
            os.mkdir(dirname)
        self.filename = f'{name}.xlsx'
        self.headers = FileWriterXL.setHeaders(fieldnames)
        self.wb = Workbook() # create new workbook
        self.ws = self.wb.active
        self.setWorksheet() # define the headers and worksheet properties

    def setHeaders(fieldnames):
        '''
        @param fieldnames - iterable list of fieldnames
        '''
        headers = {}
        for itx, fieldname in enumerate(fieldnames):
            headers[fieldname] = itx+1
        return headers

    def setWorksheet(self):
        self.ws.title = 'DATA'
        self.ws.sheet_properties.tabColor = 'FF69B4'

        for header, pos in self.headers.items():
            self.ws.cell(row=1, column=pos, value=header)
        self.currentRow = 2 # set the last completed row

    def bulkWrite(self, entries):
        '''
        @param entries - array of dictionary entries with key's containing the col headers
        '''
        for entry in entries:
            for key, value in entry.items():
                self.ws.cell(row=self.currentRow, column=self.headers[key], value=value)
            self.currentRow += 1 # increment row
        return

    def bulkWriteSheet(self, sheetname, entries):
        '''
        @param entries - array of dictionary entries with key's containing the col headers
        '''

        ws = self.wb.create_sheet(title=sheetname)
        currentRow = 1 # set the current row as the header

        # Write the header
        for key, value in self.headers.items():
            ws.cell(row=currentRow, column=value, value=key)
        currentRow += 1

        for entry in entries:
            for key, value in entry.items():
                ws.cell(row=currentRow, column=self.headers[key], value=value)
            currentRow += 1 # increment row
        return

    def singleWrite(self, entry):
        '''
        @param entry - dictionary entry
        '''
        for key, value in entry.items():
            self.ws.cell(row=self.currentRow, column=self.headers[key], value=value)
        self.currentRow += 1 # increment row
        return

    def printHeaders(self):
        print(self.headers.keys())

    def getFilename(self):
        return self.filename

    def close(self):
        '''
        Saves the excel file to file and then closes the file properly.
        '''
        return self.wb.save(self.filename)
